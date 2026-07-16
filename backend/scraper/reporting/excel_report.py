"""
excel_report.py

Writes the daily_report.xlsx workbook with two sheets:
  1. "Daily Report" — same rows as daily_report.csv
  2. "Summary"       — execution summary stats

Uses openpyxl directly (no pandas dependency required for this module).
"""

import logging
from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from reporting.csv_report import CSV_COLUMNS

logger = logging.getLogger(__name__)

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

SUMMARY_ROWS: List[str] = [
    "Execution Date",
    "Start Time",
    "End Time",
    "Duration",
    "Total Documents Found",
    "New Documents",
    "Duplicate Documents",
    "Failed Documents",
    "Skipped Documents",
    "Files Uploaded to S3",
    "Metadata Updated",
    "Next Scheduled Run",
]


def _write_header(ws, columns: List[str]) -> None:
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for col_idx, col_name in enumerate(columns, start=1):
        width = max(14, min(45, len(col_name) + 4))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_excel(
    records: List[Dict],
    summary: Dict,
    connector_summary: Dict[str, Dict],
    output_path: Path,
) -> Path:
    """
    Write the full daily_report.xlsx with a "Daily Report" sheet and a
    "Summary" sheet (including a connector-wise breakdown table).
    """

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # ------------------------------------------------------------------
    # Sheet 1: Daily Report
    # ------------------------------------------------------------------
    ws_report = wb.active
    ws_report.title = "Daily Report"
    _write_header(ws_report, CSV_COLUMNS)

    for row_idx, record in enumerate(records, start=2):
        for col_idx, col_name in enumerate(CSV_COLUMNS, start=1):
            ws_report.cell(row=row_idx, column=col_idx, value=record.get(col_name, ""))

    ws_report.freeze_panes = "A2"

    # ------------------------------------------------------------------
    # Sheet 2: Summary
    # ------------------------------------------------------------------
    ws_summary = wb.create_sheet("Summary")
    _write_header(ws_summary, ["Metric", "Value"])

    row = 2
    for label in SUMMARY_ROWS:
        ws_summary.cell(row=row, column=1, value=label)
        ws_summary.cell(row=row, column=2, value=summary.get(label, ""))
        row += 1

    # Connector-wise summary table
    row += 1
    ws_summary.cell(row=row, column=1, value="Connector-wise Summary").font = Font(bold=True)
    row += 1

    conn_headers = ["Connector", "Total", "New", "Duplicate", "Failed", "Duration (s)"]
    for col_idx, header in enumerate(conn_headers, start=1):
        cell = ws_summary.cell(row=row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    row += 1

    for connector_name, stats in connector_summary.items():
        ws_summary.cell(row=row, column=1, value=connector_name)
        ws_summary.cell(row=row, column=2, value=stats.get("total", 0))
        ws_summary.cell(row=row, column=3, value=stats.get("new", 0))
        ws_summary.cell(row=row, column=4, value=stats.get("duplicate", 0))
        ws_summary.cell(row=row, column=5, value=stats.get("failed", 0))
        ws_summary.cell(row=row, column=6, value=stats.get("duration_seconds", 0))
        row += 1

    ws_summary.column_dimensions["A"].width = 28
    ws_summary.column_dimensions["B"].width = 22

    wb.save(output_path)

    logger.info(f"Excel report written: {output_path} ({len(records)} rows)")
    return output_path
